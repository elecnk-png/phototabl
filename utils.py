import os
from datetime import datetime
from typing import List
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
import logging
from PIL import Image

logger = logging.getLogger(__name__)

def create_excel_with_images(entries: List[dict], user_id: int) -> str:
    """
    Создает Excel файл с записями и встроенными изображениями
    Returns:
        путь к созданному файлу
    """
    try:
        # Создаем директорию для экспорта, если её нет
        os.makedirs('exports', exist_ok=True)
        
        # Создаем новый Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Записи с фото"
        
        # Заголовки
        headers = ['№', 'Название', 'Описание', 'Фото', 'Количество фото', 'Дата создания']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 5   # №
        ws.column_dimensions['B'].width = 30  # Название
        ws.column_dimensions['C'].width = 40  # Описание
        ws.column_dimensions['D'].width = 25  # Фото
        ws.column_dimensions['E'].width = 15  # Количество фото
        ws.column_dimensions['F'].width = 20  # Дата создания
        
        # Заполняем данными
        current_row = 2
        for idx, entry in enumerate(entries, 1):
            logger.info(f"Обработка записи {idx}: {entry.get('name', 'Без названия')}")
            
            # №
            ws.cell(row=current_row, column=1, value=idx)
            
            # Название
            name_cell = ws.cell(row=current_row, column=2, value=entry.get('name', ''))
            name_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Описание
            desc_cell = ws.cell(row=current_row, column=3, value=entry.get('description', ''))
            desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Количество фото
            photos = entry.get('photos', [])
            ws.cell(row=current_row, column=5, value=len(photos))
            
            # Дата создания
            timestamp = entry.get('timestamp', '')
            if timestamp:
                try:
                    dt = datetime.fromisoformat(timestamp)
                    ws.cell(row=current_row, column=6, value=dt.strftime('%d.%m.%Y %H:%M'))
                except:
                    ws.cell(row=current_row, column=6, value=timestamp)
            
            # Добавляем изображения в ячейку Фото
            if photos:
                # Фильтруем существующие фото
                existing_photos = []
                for photo_path in photos[:3]:  # Максимум 3 фото на запись
                    if os.path.exists(photo_path):
                        existing_photos.append(photo_path)
                    else:
                        logger.warning(f"Фото не найдено: {photo_path}")
                
                if existing_photos:
                    # Создаем временное составное изображение
                    combined_image = create_photo_collage(existing_photos)
                    if combined_image:
                        # Сохраняем временное изображение
                        temp_img_path = f"exports/temp_collage_{user_id}_{idx}_{datetime.now().timestamp()}.png"
                        combined_image.save(temp_img_path, format='PNG')
                        
                        # Вставляем в Excel
                        img = XLImage(temp_img_path)
                        
                        # Масштабируем изображение
                        img.width = 200
                        img.height = 150
                        
                        # Вставляем изображение
                        cell_address = f'D{current_row}'
                        ws.add_image(img, cell_address)
                        
                        # Устанавливаем высоту строки
                        ws.row_dimensions[current_row].height = 100
                        
                        # Удаляем временный файл
                        try:
                            os.remove(temp_img_path)
                        except Exception as e:
                            logger.error(f"Ошибка при удалении временного файла {temp_img_path}: {e}")
                    else:
                        ws.cell(row=current_row, column=4, value="[Ошибка создания коллажа]")
                else:
                    ws.cell(row=current_row, column=4, value="[Фото не найдены]")
            else:
                ws.cell(row=current_row, column=4, value="Нет фото")
            
            current_row += 1
        
        # Добавляем границы для всей таблицы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
        
        # Сохраняем файл
        filename = f"export_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('exports', filename)
        wb.save(filepath)
        
        logger.info(f"✅ Excel файл с изображениями создан: {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"❌ Ошибка при создании Excel с изображениями: {e}")
        # Пробуем создать простой Excel без изображений
        return create_simple_excel(entries, user_id)

def create_photo_collage(photo_paths: List[str], max_width=600, max_height=400) -> Image.Image:
    """
    Создает коллаж из нескольких фотографий
    Args:
        photo_paths: список путей к фото
        max_width: максимальная ширина коллажа
        max_height: максимальная высота коллажа
    Returns:
        Image объект с коллажем или None в случае ошибки
    """
    try:
        images = []
        for path in photo_paths:
            if os.path.exists(path):
                try:
                    img = Image.open(path)
                    # Конвертируем в RGB если нужно
                    if img.mode in ('RGBA', 'LA', 'P'):
                        # Создаем белый фон для прозрачных изображений
                        if img.mode == 'RGBA':
                            background = Image.new('RGB', img.size, (255, 255, 255))
                            background.paste(img, mask=img.split()[3])  # Используем альфа-канал как маску
                            img = background
                        else:
                            img = img.convert('RGB')
                    images.append(img)
                except Exception as e:
                    logger.error(f"Ошибка при открытии фото {path}: {e}")
            else:
                logger.warning(f"Файл не существует: {path}")
        
        if not images:
            logger.warning("Нет доступных изображений для создания коллажа")
            return None
        
        if len(images) == 1:
            # Одно изображение - просто изменяем размер
            img = images[0]
            # Сохраняем пропорции
            img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
            return img
        
        # Несколько изображений - создаем коллаж
        if len(images) == 2:
            # Два изображения - горизонтально
            collage = Image.new('RGB', (max_width, max_height), 'white')
            
            # Рассчитываем ширину для каждого изображения с сохранением пропорций
            heights = []
            widths = []
            for img in images:
                aspect = img.width / img.height
                width = int(max_height * aspect)
                if width > max_width // 2:
                    width = max_width // 2
                heights.append(max_height)
                widths.append(width)
            
            # Изменяем размер и вставляем
            x = 0
            for i, img in enumerate(images):
                img_resized = img.copy()
                img_resized.thumbnail((widths[i], heights[i]), Image.Resampling.LANCZOS)
                y = (max_height - img_resized.height) // 2
                collage.paste(img_resized, (x, y))
                x += widths[i]
            
            return collage
        
        elif len(images) == 3:
            # Три изображения - большое слева, два маленьких справа
            collage = Image.new('RGB', (max_width, max_height), 'white')
            
            # Левое изображение (50% ширины)
            left_width = max_width // 2
            img_left = images[0].copy()
            img_left.thumbnail((left_width, max_height), Image.Resampling.LANCZOS)
            y_left = (max_height - img_left.height) // 2
            collage.paste(img_left, (0, y_left))
            
            # Правая колонка (два изображения)
            right_width = max_width - left_width
            right_height = max_height // 2
            
            for i in range(1, min(3, len(images))):
                img_right = images[i].copy()
                img_right.thumbnail((right_width, right_height), Image.Resampling.LANCZOS)
                x = left_width + (right_width - img_right.width) // 2
                y = (i-1) * right_height + (right_height - img_right.height) // 2
                collage.paste(img_right, (x, y))
            
            return collage
        
        else:
            # 4+ изображений - сетка 2x2
            grid_size = 2
            cell_width = max_width // grid_size
            cell_height = max_height // grid_size
            
            collage = Image.new('RGB', (max_width, max_height), 'white')
            
            for i, img in enumerate(images[:4]):  # Максимум 4 фото
                if i >= grid_size * grid_size:
                    break
                
                img_resized = img.copy()
                img_resized.thumbnail((cell_width, cell_height), Image.Resampling.LANCZOS)
                
                x = (i % grid_size) * cell_width + (cell_width - img_resized.width) // 2
                y = (i // grid_size) * cell_height + (cell_height - img_resized.height) // 2
                
                collage.paste(img_resized, (x, y))
            
            return collage
        
    except Exception as e:
        logger.error(f"Ошибка при создании коллажа: {e}")
        return None

def create_simple_excel(entries: List[dict], user_id: int) -> str:
    """
    Создает простой Excel файл с записями (без изображений) как запасной вариант
    """
    try:
        os.makedirs('exports', exist_ok=True)
        
        # Подготавливаем данные
        data = []
        for entry in entries:
            # Формируем пути к фото
            photo_paths = entry.get('photos', [])
            photo_info = []
            for path in photo_paths:
                if os.path.exists(path):
                    photo_info.append(os.path.basename(path))
                else:
                    photo_info.append("[файл не найден]")
            
            data.append({
                'Название': entry.get('name', ''),
                'Описание': entry.get('description', ''),
                'Количество фото': len(photo_paths),
                'Файлы фото': ', '.join(photo_info) if photo_info else 'Нет фото',
                'Дата создания': entry.get('timestamp', '')
            })
        
        df = pd.DataFrame(data)
        
        filename = f"export_simple_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('exports', filename)
        
        # Создаем Excel с настройками
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Записи', index=False)
            
            # Настраиваем ширину колонок
            worksheet = writer.sheets['Записи']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"✅ Простой Excel файл создан: {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"❌ Ошибка при создании простого Excel: {e}")
        raise

def format_entry_preview(entry: dict) -> str:
    """Форматирует запись для предпросмотра в Telegram"""
    lines = [
        f"📝 **{entry.get('name', 'Без названия')}**",
        f"📋 {entry.get('description', 'Нет описания')}",
        f"📸 Фото: {len(entry.get('photos', []))} шт.",
    ]
    
    if 'timestamp' in entry:
        try:
            dt = datetime.fromisoformat(entry['timestamp'])
            lines.append(f"🕐 {dt.strftime('%d.%m.%Y %H:%M')}")
        except:
            lines.append(f"🕐 {entry['timestamp']}")
    
    return '\n'.join(lines)

def validate_photo(file_path: str) -> bool:
    """Проверяет, является ли файл корректным изображением"""
    try:
        if not os.path.exists(file_path):
            logger.warning(f"Файл не существует: {file_path}")
            return False
        
        # Проверяем размер файла (не больше 20MB)
        file_size = os.path.getsize(file_path)
        if file_size > 20 * 1024 * 1024:
            logger.warning(f"Файл слишком большой: {file_size} байт")
            return False
        
        # Проверяем что это изображение
        with Image.open(file_path) as img:
            img.verify()
        
        # Дополнительная проверка - пробуем открыть после verify
        with Image.open(file_path) as img:
            img.load()
        
        logger.info(f"✅ Фото успешно проверено: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Ошибка валидации фото {file_path}: {e}")
        return False

def cleanup_old_files(days: int = 7):
    """Очищает старые временные файлы"""
    now = datetime.now().timestamp()
    deleted_count = 0
    error_count = 0
    
    for directory in ['photos', 'exports']:
        if os.path.exists(directory):
            for filename in os.listdir(directory):
                filepath = os.path.join(directory, filename)
                if os.path.isfile(filepath):
                    try:
                        file_time = os.path.getmtime(filepath)
                        if now - file_time > days * 24 * 60 * 60:
                            os.remove(filepath)
                            deleted_count += 1
                            logger.info(f"Удален старый файл: {filepath}")
                    except Exception as e:
                        error_count += 1
                        logger.error(f"Ошибка при удалении {filepath}: {e}")
    
    logger.info(f"Очистка завершена. Удалено файлов: {deleted_count}, ошибок: {error_count}")

def get_photo_stats(entries: List[dict]) -> dict:
    """Получает статистику по фото для всех записей"""
    total_photos = 0
    entries_with_photos = 0
    photo_sizes = []
    
    for entry in entries:
        photos = entry.get('photos', [])
        if photos:
            entries_with_photos += 1
            total_photos += len(photos)
            
            # Собираем информацию о размерах фото
            for photo_path in photos:
                if os.path.exists(photo_path):
                    try:
                        size = os.path.getsize(photo_path)
                        photo_sizes.append(size)
                    except:
                        pass
    
    return {
        'total_entries': len(entries),
        'entries_with_photos': entries_with_photos,
        'total_photos': total_photos,
        'avg_photos_per_entry': total_photos / len(entries) if entries else 0,
        'avg_photo_size_mb': sum(photo_sizes) / len(photo_sizes) / (1024*1024) if photo_sizes else 0
    }
